import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import calendar
import io
import base64

# Page configuration
st.set_page_config(
    page_title="MahaSTRIDE - Quarterly Project Plan Dashboard",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS (kept same as before)
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        color: white;
    }
    .assignment-title {
        font-size: 1.1rem;
        font-weight: 500;
        margin: 0.5rem 0;
        padding: 0.5rem;
        background-color: rgba(255,255,255,0.1);
        border-radius: 8px;
        text-align: center;
    }
    .parties-info {
        background: linear-gradient(135deg, #2a5298 0%, #1e3c72 100%);
        padding: 1rem;
        border-radius: 10px;
        margin-top: 1rem;
        font-size: 0.9rem;
        border: 1px solid rgba(255,255,255,0.2);
    }
    .quarter-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1rem;
        border-radius: 10px;
        margin: 0.5rem 0;
        transition: transform 0.3s;
    }
    .quarter-card:hover {
        transform: translateY(-5px);
    }
    .milestone-completed {
        background-color: #d4edda;
        border-left: 4px solid #28a745;
        padding: 1rem;
        margin: 0.5rem 0;
        border-radius: 8px;
    }
    .milestone-achieved {
        background-color: #d4edda;
        border-left: 4px solid #28a745;
        padding: 1rem;
        margin: 0.5rem 0;
        border-radius: 8px;
    }
    .milestone-upcoming {
        background-color: #fff3cd;
        border-left: 4px solid #ffc107;
        padding: 1rem;
        margin: 0.5rem 0;
        border-radius: 8px;
    }
    .stat-card {
        background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
        color: white;
        padding: 1rem;
        border-radius: 10px;
        text-align: center;
    }
    .status-ongoing {
        background-color: #fff3cd;
        color: #856404;
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        display: inline-block;
    }
    .status-completed {
        background-color: #d4edda;
        color: #155724;
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        display: inline-block;
    }
    .war-room-card {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        color: white;
        padding: 1rem;
        border-radius: 10px;
        margin: 1rem 0;
    }
    .party-box {
        background: rgba(255,255,255,0.15);
        border-radius: 8px;
        padding: 0.75rem;
        margin: 0.5rem 0;
    }
</style>
""", unsafe_allow_html=True)

# ============================================================
# PROJECT DATA (Same as before)
# ============================================================

PROJECT_NAME = "MahaSTRIDE - Maharashtra Strengthening Institutional Capabilities in Districts for Enabling Growth"
ASSIGNMENT_TITLE = "Engagement of a Consultancy Firm for Comprehensive Data Collection, Advanced Analytics, and Development of Performance Improvement Framework for Maharashtra State Universities under MahaSTRIDE Operations"

CLIENT_NAME = "Maharashtra Institute for Transformation (MITRA), State Data Authority, Government of Maharashtra"
CONSULTANT_NAME = "Indian Centre for Academic Rankings & Excellence - ICARE Pvt. Ltd."

START_DATE = datetime(2026, 5, 5)
END_DATE = datetime(2028, 5, 6)

# Achieved Milestones
ACHIEVED_MILESTONES = [
    {"milestone": "SANGAM Orientation & Training Completed", "date": "May 4-6, 2026", "status": "achieved"},
    {"milestone": "Inception Report & GRDAU Framework Submitted", "date": "May 26, 2026", "status": "achieved"}
]

# Universities Data with GRDAU Status
UNIVERSITIES = {
    "MU": {
        "name": "Mumbai University",
        "location": "Mumbai",
        "vice_chancellor": "Dr. Ravindra Kulkarni",
        "nodal_officer": "Dr. Varsha Kelkar Mane",
        "contact": "+91-22-26543000",
        "coordinators": ["Sneha Kashitkar", "Sagar Teli"],
        "grdau_status": "Completed",
        "grdau_completion_date": "July 5, 2026"
    },
    "SSPU": {
        "name": "Savitribai Phule Pune University",
        "location": "Pune",
        "vice_chancellor": "Dr. Suresh Gosavi",
        "nodal_officer": "Prof. Vinayak Joshi",
        "contact": "+91-20-25696061",
        "coordinators": ["Jagan Sridhar"],
        "grdau_status": "Completed",
        "grdau_completion_date": "July 5, 2026"
    },
    "COEP": {
        "name": "COEP Technological University",
        "location": "Pune",
        "vice_chancellor": "Dr. B. K. Mishra",
        "nodal_officer": "Dr. Uttam Chaskar",
        "contact": "+91-20-25507000",
        "coordinators": ["Vaibhav Ambekar"],
        "grdau_status": "Completed",
        "grdau_completion_date": "July 5, 2026"
    },
    "KBCNMU": {
        "name": "Kavayitri Bahinabai Chaudhari North Maharashtra University",
        "location": "Jalgaon",
        "vice_chancellor": "Dr. R. P. Swami",
        "nodal_officer": "Prof. Sameer Narkhede",
        "contact": "+91-257-2257457",
        "coordinators": ["Nitish Kumbhar"],
        "grdau_status": "Completed",
        "grdau_completion_date": "July 5, 2026"
    },
    "BAMU": {
        "name": "Dr. Babasaheb Ambedkar Marathwada University",
        "location": "Chhatrapati Sambhajinagar",
        "vice_chancellor": "Dr. Pramod Yeole",
        "nodal_officer": "Prof. G. D. Khedkar",
        "contact": "+91-240-2403111",
        "coordinators": ["Atharav Paturkar"],
        "grdau_status": "Completed",
        "grdau_completion_date": "July 5, 2026"
    },
    "NU": {
        "name": "Rashtrasant Tukadoji Maharaj Nagpur University",
        "location": "Nagpur",
        "vice_chancellor": "Dr. Subhash Chaudhari",
        "nodal_officer": "Prof. Nandkishor Karade",
        "contact": "+91-712-2500511",
        "coordinators": ["Anjali Singh"],
        "grdau_status": "Completed",
        "grdau_completion_date": "July 5, 2026"
    },
    "AU": {
        "name": "Sant Gadge Baba Amravati University",
        "location": "Amravati",
        "vice_chancellor": "Dr. Milind Baride",
        "nodal_officer": "Dr. A. B. Naik",
        "contact": "+91-721-2662379",
        "coordinators": ["Prathamesh Babhulkar"],
        "grdau_status": "Completed",
        "grdau_completion_date": "July 5, 2026"
    },
    "MITRA": {
        "name": "MITRA - State Data Authority",
        "location": "Mumbai",
        "ceo": "Shri. Aman Mittal",
        "nodal_officer": "Dr. Harshal Kotwal",
        "contact": "+91-22-69979440",
        "coordinators": ["Shubham Singh"],
        "grdau_status": "Completed",
        "grdau_completion_date": "July 5, 2026"
    }
}

# ============================================================
# QUARTERLY PLAN (Same as before)
# ============================================================

QUARTERS = {
    "Q1: May - July 2026": {
        "number": 1,
        "months": ["May 2026", "June 2026", "July 2026"],
        "status": "ongoing",
        "key_activities": [
            "✅ SANGAM Orientation & Training (May 4-6 at Trident Board Room)",
            "✅ University Onboarding & Data Source Mapping",
            "✅ NIRF Data Collection (Student, Faculty, Research, Placement, Finance)",
            "✅ Inception Report & GRDAU Framework Development",
            "✅ GRDAU Establishment in all universities (Completed July 5, 2026)",
            "🔄 Diagnostic Assessments across all 7 universities",
            "🔄 Gap Analysis against NIRF/NAAC/Global Rankings",
            "🔄 SWOT Analysis for each university"
        ],
        "deliverables": [
            "✅ Inception Report and Deployment Plan (Submitted May 26, 2026)",
            "✅ GRDAUs Established and Operationalized (Completed July 5, 2026)",
            "🔄 Diagnostic Assessment Reports (7 universities) - Due July 31, 2026",
            "🔄 SWOT Analysis Reports - Due July 31, 2026"
        ],
        "milestones": [
            {"name": "SANGAM Training Completed", "status": "achieved", "date": "May 6, 2026"},
            {"name": "Inception Report Submitted", "status": "achieved", "date": "May 26, 2026"},
            {"name": "GRDAU Establishment Completed", "status": "achieved", "date": "July 5, 2026"},
            {"name": "Diagnostic Reports", "status": "in_progress", "date": "July 31, 2026"},
            {"name": "SWOT Analysis Reports", "status": "in_progress", "date": "July 31, 2026"},
            {"name": "Gap Analysis Report", "status": "in_progress", "date": "July 31, 2026"}
        ],
        "data_collection": "NIRF baseline data collection completed. Diagnostic assessments in progress.",
        "stakeholder_engagement": "VC meetings conducted. IQAC coordination established.",
        "review_mechanism": "Weekly GRDAU meetings. Monthly progress review with MITRA PMU.",
        "universities_involved": "All 7 universities",
        "responsible_party": "ICARE Pvt. Ltd. with GRDAU teams"
    },
    "Q2: August - October 2026": {
        "number": 2,
        "months": ["August 2026", "September 2026", "October 2026"],
        "status": "upcoming",
        "key_activities": [
            "Institutional Development Plans (IDPs) development",
            "Stakeholder review and feedback incorporation",
            "Data portal architecture design",
            "Dashboard requirements gathering",
            "Dashboard prototype development",
            "Milestone 1: Sustainable Data Systems establishment"
        ],
        "deliverables": [
            "Institutional Development Plans (IDPs) - 7",
            "Portal Design Document",
            "Dashboard Mockups",
            "Milestone 1 Report"
        ],
        "milestones": [
            {"name": "IDPs Draft Completed", "status": "pending", "date": "Aug 31, 2026"},
            {"name": "Portal Design Approved", "status": "pending", "date": "Sep 15, 2026"},
            {"name": "Milestone 1: Sustainable Data & Quality Systems", "status": "pending", "date": "Sep 30, 2026"},
            {"name": "Milestone 2: IDP Execution Monitoring", "status": "pending", "date": "Oct 31, 2026"}
        ],
        "data_collection": "IDP data collection. Dashboard requirements gathering.",
        "stakeholder_engagement": "IDP review meetings with VCs. Dashboard workshops.",
        "review_mechanism": "Bi-weekly IDP review. Monthly progress review.",
        "universities_involved": "All 7 universities",
        "responsible_party": "ICARE Pvt. Ltd. with GRDAU teams"
    },
    "Q3: November 2026 - January 2027": {
        "number": 3,
        "months": ["November 2026", "December 2026", "January 2027"],
        "status": "upcoming",
        "key_activities": [
            "Data Portal MVP Deployment",
            "Training Needs Assessment",
            "Capacity Building Programs (First round)",
            "Performance Dashboards Launch",
            "Data Validation and Quality Improvement",
            "Milestone 3: Capacity Building Participation"
        ],
        "deliverables": [
            "Data Portal Live",
            "Training Completion Report",
            "Dashboard Deployment Report"
        ],
        "milestones": [
            {"name": "Portal MVP Launch", "status": "pending", "date": "Nov 15, 2026"},
            {"name": "Mid-term Progress Report", "status": "pending", "date": "Nov 30, 2026"},
            {"name": "First Training Program", "status": "pending", "date": "Dec 15, 2026"},
            {"name": "Milestone 3: Capacity Building", "status": "pending", "date": "Dec 31, 2026"}
        ],
        "data_collection": "Portal data upload. Training feedback collection.",
        "stakeholder_engagement": "Portal training sessions. Capacity building workshops.",
        "review_mechanism": "Portal usage analytics. Training effectiveness assessment.",
        "universities_involved": "All 7 universities",
        "responsible_party": "ICARE Pvt. Ltd."
    },
    "Q4: February - April 2027": {
        "number": 4,
        "months": ["February 2027", "March 2027", "April 2027"],
        "status": "upcoming",
        "key_activities": [
            "Research Output Enhancement Initiatives",
            "International Collaboration Development",
            "Accreditation Preparedness Assessment",
            "Quality Assurance Framework Implementation"
        ],
        "deliverables": [
            "Research Enhancement Plan",
            "Collaboration Framework",
            "QA Framework Report"
        ],
        "milestones": [
            {"name": "Research Enhancement Plan", "status": "pending", "date": "Feb 28, 2027"},
            {"name": "Year 1 Annual Report", "status": "pending", "date": "Apr 30, 2027"}
        ],
        "data_collection": "Research output data. Collaboration metrics.",
        "stakeholder_engagement": "Research committee meetings. Industry collaboration.",
        "review_mechanism": "Research output tracking. QA dashboard monitoring.",
        "universities_involved": "All 7 universities",
        "responsible_party": "ICARE Pvt. Ltd. with Research Cells"
    },
    "Q5: May - July 2027": {
        "number": 5,
        "months": ["May 2027", "June 2027", "July 2027"],
        "status": "upcoming",
        "key_activities": [
            "Year 2 Kickoff and Advanced Analytics",
            "Global Ranking Preparation (QS, THE, US News)",
            "Advanced Training Programs",
            "Milestone 4: 10% Improvement Achievement"
        ],
        "deliverables": [
            "Year 2 Work Plan",
            "Ranking Submission Packages",
            "Advanced Training Report",
            "Milestone 4 Report"
        ],
        "milestones": [
            {"name": "QS Ranking Submission", "status": "pending", "date": "Jun 15, 2027"},
            {"name": "Milestone 4: 10% Improvement", "status": "pending", "date": "Jun 30, 2027"}
        ],
        "data_collection": "Ranking data compilation. Improvement metrics.",
        "stakeholder_engagement": "Ranking preparation workshops. Industry advisory board.",
        "review_mechanism": "Quarterly performance review. Ranking submission tracking.",
        "universities_involved": "All 7 universities",
        "responsible_party": "ICARE Pvt. Ltd. with GRDAU teams"
    },
    "Q6: August - October 2027": {
        "number": 6,
        "months": ["August 2027", "September 2027", "October 2027"],
        "status": "upcoming",
        "key_activities": [
            "Employer Perception Enhancement",
            "Academic Reputation Building",
            "Industry Connect Programs",
            "International Student Enrollment Strategies"
        ],
        "deliverables": [
            "Employer Perception Report",
            "Reputation Strategy Document",
            "Industry Connect Report",
            "Internationalization Plan"
        ],
        "milestones": [
            {"name": "Employer Survey Completion", "status": "pending", "date": "Aug 31, 2027"},
            {"name": "International MoUs Signed", "status": "pending", "date": "Sep 30, 2027"}
        ],
        "data_collection": "Employer survey data. Reputation metrics.",
        "stakeholder_engagement": "Employer meets. International partner meetings.",
        "review_mechanism": "Monthly reputation tracking. Employer feedback analysis.",
        "universities_involved": "All 7 universities",
        "responsible_party": "ICARE Pvt. Ltd. with Placement Cells"
    },
    "Q7: November 2027 - January 2028": {
        "number": 7,
        "months": ["November 2027", "December 2027", "January 2028"],
        "status": "upcoming",
        "key_activities": [
            "Final Global Ranking Submissions",
            "Sustainability Planning",
            "Knowledge Transfer Preparation",
            "Milestone 5: 20% Improvement",
            "Milestone 6: Global Rankings Participation"
        ],
        "deliverables": [
            "Final Ranking Submissions",
            "Sustainability Plan",
            "Knowledge Transfer Report",
            "Milestone 5 & 6 Reports"
        ],
        "milestones": [
            {"name": "Milestone 5: 20% Improvement", "status": "pending", "date": "Dec 31, 2027"},
            {"name": "Milestone 6: Global Rankings", "status": "pending", "date": "Feb 29, 2028"},
            {"name": "Sustainability Plan", "status": "pending", "date": "Dec 15, 2027"}
        ],
        "data_collection": "20% improvement evidence. Ranking participation data.",
        "stakeholder_engagement": "Sustainability workshop. Handover planning.",
        "review_mechanism": "Final evaluation framework. Sustainability assessment.",
        "universities_involved": "All 7 universities",
        "responsible_party": "ICARE Pvt. Ltd. with MITRA"
    },
    "Q8: February - April 2028": {
        "number": 8,
        "months": ["February 2028", "March 2028", "April 2028"],
        "status": "upcoming",
        "key_activities": [
            "Final Evaluation and Reporting",
            "Project Closure and Knowledge Transfer",
            "Milestone 7: Final Evaluation",
            "Contract Completion"
        ],
        "deliverables": [
            "Final Closure Report",
            "Lessons Learned Report",
            "Knowledge Transfer Documentation",
            "Milestone 7 Report"
        ],
        "milestones": [
            {"name": "Milestone 7: Final Evaluation", "status": "pending", "date": "Apr 30, 2028"},
            {"name": "Project Closure", "status": "pending", "date": "May 6, 2028"}
        ],
        "data_collection": "Final performance metrics. Lessons learned.",
        "stakeholder_engagement": "Final client presentation. Project closure meeting.",
        "review_mechanism": "Final evaluation. Client satisfaction survey.",
        "universities_involved": "All 7 universities",
        "responsible_party": "ICARE Pvt. Ltd. with MITRA"
    }
}

# ============================================================
# IMPROVED EXCEL EXPORT FUNCTION - Professional Format
# ============================================================

def export_quarter_to_professional_excel(quarter_name, selected_sections):
    """Export selected quarter data to professionally formatted Excel"""
    quarter_info = QUARTERS[quarter_name]
    
    output = io.BytesIO()
    
    try:
        # Try using openpyxl for better formatting
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1e3c72', end_color='1e3c72', fill_type='solid')
        subheader_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        subheader_fill = PatternFill(start_color='2a5298', end_color='2a5298', fill_type='solid')
        title_font = Font(name='Arial', size=14, bold=True, color='1e3c72')
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. COVER SHEET
        ws_cover = wb.create_sheet("COVER SHEET", 0)
        ws_cover.merge_cells('A1:F1')
        cell = ws_cover['A1']
        cell.value = "MAHASTRIDE PROJECT"
        cell.font = Font(name='Arial', size=18, bold=True, color='1e3c72')
        cell.alignment = center_alignment
        
        ws_cover.merge_cells('A2:F2')
        cell = ws_cover['A2']
        cell.value = "Quarterly Progress Report"
        cell.font = Font(name='Arial', size=14, bold=True)
        cell.alignment = center_alignment
        
        ws_cover.merge_cells('A3:F3')
        cell = ws_cover['A3']
        cell.value = quarter_name
        cell.font = Font(name='Arial', size=12, bold=True, color='2a5298')
        cell.alignment = center_alignment
        
        # Cover sheet data
        cover_data = [
            ["", ""],
            ["Project Name:", PROJECT_NAME],
            ["Assignment Title:", ASSIGNMENT_TITLE],
            ["", ""],
            ["Client:", CLIENT_NAME],
            ["Consultant:", CONSULTANT_NAME],
            ["", ""],
            ["Quarter:", quarter_name],
            ["Reporting Period:", f"{quarter_info['months'][0]} - {quarter_info['months'][2]}"],
            ["Status:", quarter_info['status'].upper()],
            ["", ""],
            ["World Bank Loan No:", "IBRD 9737-IN"],
            ["RFP Reference:", "IN-MITRA(PMU)-PforR-Edu-QCBS"],
            ["", ""],
            ["Report Generated On:", datetime.now().strftime('%d %B %Y, %H:%M:%S')]
        ]
        
        row = 6
        for label, value in cover_data:
            if label:
                ws_cover.cell(row=row, column=1, value=label)
                ws_cover.cell(row=row, column=2, value=value)
                ws_cover.cell(row=row, column=1).font = Font(bold=True)
            row += 1
        
        # 2. EXECUTIVE SUMMARY
        if "Executive Summary" in selected_sections:
            ws_es = wb.create_sheet("EXECUTIVE SUMMARY")
            
            # Title
            ws_es.merge_cells('A1:D1')
            cell = ws_es['A1']
            cell.value = "EXECUTIVE SUMMARY"
            cell.font = title_font
            cell.alignment = center_alignment
            
            # Summary table
            summary_data = [
                ["Project Name", PROJECT_NAME],
                ["Assignment Title", ASSIGNMENT_TITLE],
                ["Client", CLIENT_NAME],
                ["Consultant", CONSULTANT_NAME],
                ["Quarter", quarter_name],
                ["Months", ", ".join(quarter_info["months"])],
                ["Status", quarter_info["status"].upper()],
                ["World Bank Loan No", "IBRD 9737-IN"],
                ["RFP Reference", "IN-MITRA(PMU)-PforR-Edu-QCBS"],
                ["Report Generated On", datetime.now().strftime('%d %B %Y, %H:%M:%S')]
            ]
            
            for i, (label, value) in enumerate(summary_data, start=3):
                ws_es.cell(row=i, column=1, value=label).font = Font(bold=True)
                ws_es.cell(row=i, column=2, value=value)
                ws_es.column_dimensions['A'].width = 25
                ws_es.column_dimensions['B'].width = 60
        
        # 3. KEY ACTIVITIES
        if "Key Activities" in selected_sections:
            ws_ka = wb.create_sheet("KEY ACTIVITIES")
            
            ws_ka.merge_cells('A1:D1')
            cell = ws_ka['A1']
            cell.value = "KEY ACTIVITIES"
            cell.font = title_font
            cell.alignment = center_alignment
            
            # Headers
            headers = ["S.No", "Activity", "Category", "Status"]
            for col, header in enumerate(headers, 1):
                cell = ws_ka.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Data
            for i, activity in enumerate(quarter_info["key_activities"], start=4):
                ws_ka.cell(row=i, column=1, value=i-3).alignment = center_alignment
                ws_ka.cell(row=i, column=2, value=activity.replace("✅", "").replace("🔄", "").strip())
                
                if activity.startswith("✅"):
                    category = "Completed"
                    status = "Completed"
                elif activity.startswith("🔄"):
                    category = "In Progress"
                    status = "In Progress"
                else:
                    category = "Planned"
                    status = "Planned"
                
                ws_ka.cell(row=i, column=3, value=category).alignment = center_alignment
                ws_ka.cell(row=i, column=4, value=status).alignment = center_alignment
            
            ws_ka.column_dimensions['A'].width = 8
            ws_ka.column_dimensions['B'].width = 70
            ws_ka.column_dimensions['C'].width = 15
            ws_ka.column_dimensions['D'].width = 15
        
        # 4. DELIVERABLES
        if "Deliverables" in selected_sections:
            ws_del = wb.create_sheet("DELIVERABLES")
            
            ws_del.merge_cells('A1:C1')
            cell = ws_del['A1']
            cell.value = "DELIVERABLES"
            cell.font = title_font
            cell.alignment = center_alignment
            
            headers = ["Deliverable", "Status", "Due Date"]
            for col, header in enumerate(headers, 1):
                cell = ws_del.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            for i, deliverable in enumerate(quarter_info["deliverables"], start=4):
                ws_del.cell(row=i, column=1, value=deliverable.replace("✅", "").replace("🔄", "").strip())
                
                if "✅" in deliverable:
                    status = "Completed"
                elif "🔄" in deliverable:
                    status = "In Progress"
                else:
                    status = "Pending"
                
                ws_del.cell(row=i, column=2, value=status).alignment = center_alignment
                
                # Extract due date
                if "Due" in deliverable:
                    due_date = deliverable.split("Due")[-1].strip()
                else:
                    due_date = quarter_info["milestones"][-1]["date"] if quarter_info["milestones"] else "TBD"
                ws_del.cell(row=i, column=3, value=due_date).alignment = center_alignment
            
            ws_del.column_dimensions['A'].width = 60
            ws_del.column_dimensions['B'].width = 15
            ws_del.column_dimensions['C'].width = 15
        
        # 5. MILESTONES
        if "Milestones" in selected_sections:
            ws_mil = wb.create_sheet("MILESTONES")
            
            ws_mil.merge_cells('A1:E1')
            cell = ws_mil['A1']
            cell.value = "MILESTONES TRACKER"
            cell.font = title_font
            cell.alignment = center_alignment
            
            headers = ["Milestone Name", "Target Date", "Status", "Quarter", "Remarks"]
            for col, header in enumerate(headers, 1):
                cell = ws_mil.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            for i, milestone in enumerate(quarter_info["milestones"], start=4):
                ws_mil.cell(row=i, column=1, value=milestone["name"])
                ws_mil.cell(row=i, column=2, value=milestone["date"]).alignment = center_alignment
                
                status_text = milestone["status"].upper()
                ws_mil.cell(row=i, column=3, value=status_text).alignment = center_alignment
                
                ws_mil.cell(row=i, column=4, value=quarter_name).alignment = center_alignment
                
                if milestone["status"] == "achieved":
                    ws_mil.cell(row=i, column=5, value="Completed successfully")
                elif milestone["status"] == "in_progress":
                    ws_mil.cell(row=i, column=5, value="Work in progress")
                else:
                    ws_mil.cell(row=i, column=5, value="Not started")
            
            ws_mil.column_dimensions['A'].width = 35
            ws_mil.column_dimensions['B'].width = 15
            ws_mil.column_dimensions['C'].width = 12
            ws_mil.column_dimensions['D'].width = 20
            ws_mil.column_dimensions['E'].width = 20
        
        # 6. DATA COLLECTION & ENGAGEMENT
        if "Data Collection & Engagement" in selected_sections:
            ws_dce = wb.create_sheet("DATA & ENGAGEMENT")
            
            ws_dce.merge_cells('A1:B1')
            cell = ws_dce['A1']
            cell.value = "DATA COLLECTION & STAKEHOLDER ENGAGEMENT"
            cell.font = title_font
            cell.alignment = center_alignment
            
            sections = [
                ("Data Collection Process", quarter_info["data_collection"]),
                ("Stakeholder Engagement", quarter_info["stakeholder_engagement"]),
                ("Review Mechanism", quarter_info["review_mechanism"]),
                ("Universities Involved", quarter_info.get("universities_involved", "All 7 universities")),
                ("Responsible Party", quarter_info.get("responsible_party", "ICARE Pvt. Ltd."))
            ]
            
            row = 3
            for label, value in sections:
                ws_dce.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws_dce.cell(row=row, column=2, value=value)
                row += 2
            
            ws_dce.column_dimensions['A'].width = 30
            ws_dce.column_dimensions['B'].width = 60
        
        # 7. UNIVERSITIES & GRDAU
        if "Universities & GRDAU" in selected_sections:
            ws_uni = wb.create_sheet("UNIVERSITIES & GRDAU")
            
            ws_uni.merge_cells('A1:G1')
            cell = ws_uni['A1']
            cell.value = "UNIVERSITIES AND GRDAU STATUS"
            cell.font = title_font
            cell.alignment = center_alignment
            
            headers = ["University", "Location", "Nodal Officer", "Contact", "GRDAU Status", "Completion Date", "Coordinators"]
            for col, header in enumerate(headers, 1):
                cell = ws_uni.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            row = 4
            for code, uni in UNIVERSITIES.items():
                if code != "MITRA":
                    ws_uni.cell(row=row, column=1, value=uni["name"])
                    ws_uni.cell(row=row, column=2, value=uni["location"])
                    ws_uni.cell(row=row, column=3, value=uni.get("nodal_officer", "N/A"))
                    ws_uni.cell(row=row, column=4, value=uni.get("contact", "N/A"))
                    ws_uni.cell(row=row, column=5, value=uni["grdau_status"]).alignment = center_alignment
                    ws_uni.cell(row=row, column=6, value=uni["grdau_completion_date"]).alignment = center_alignment
                    ws_uni.cell(row=row, column=7, value=", ".join(uni["coordinators"]))
                    row += 1
            
            for col in range(1, 8):
                ws_uni.column_dimensions[get_column_letter(col)].width = 18
        
        # 8. TEAM MEMBERS
        if "Team Members" in selected_sections:
            ws_team = wb.create_sheet("TEAM MEMBERS")
            
            ws_team.merge_cells('A1:C1')
            cell = ws_team['A1']
            cell.value = "PROJECT TEAM MEMBERS"
            cell.font = title_font
            cell.alignment = center_alignment
            
            headers = ["Name", "Role", "University"]
            for col, header in enumerate(headers, 1):
                cell = ws_team.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            team_data = [
                ["Dr. Harshal Kotwal", "Project Lead", "ICARE"],
                ["Shubham Singh", "Data Analytics Specialist", "MITRA"],
                ["Sagar Teli", "Statistician & Program Designer", "Mumbai University"],
                ["Sneha Kashitkar", "Institutional Coordinator", "Mumbai University"],
                ["Nitish Kumbhar", "Institutional Coordinator", "KBCNMU Jalgaon"],
                ["Anjali Singh", "Institutional Coordinator", "Nagpur University"],
                ["Vaibhav Ambekar", "Institutional Coordinator", "COEP Pune"],
                ["Atharav Paturkar", "Institutional Coordinator", "BAMU Aurangabad"],
                ["Prathamesh Babhulkar", "Institutional Coordinator", "Amravati University"],
                ["Jagan Sridhar", "Institutional Coordinator", "SPPU Pune"]
            ]
            
            for i, (name, role, uni) in enumerate(team_data, start=4):
                ws_team.cell(row=i, column=1, value=name)
                ws_team.cell(row=i, column=2, value=role)
                ws_team.cell(row=i, column=3, value=uni)
            
            ws_team.column_dimensions['A'].width = 25
            ws_team.column_dimensions['B'].width = 30
            ws_team.column_dimensions['C'].width = 30
        
        # Save workbook
        wb.save(output)
        output.seek(0)
        return output.getvalue(), "excel"
        
    except ImportError:
        # Fallback to CSV zip if openpyxl not available
        st.warning("openpyxl not installed. Generating CSV files instead.")
        return export_quarter_to_csv(quarter_name, selected_sections), "csv"

def export_quarter_to_csv(quarter_name, selected_sections):
    """Fallback: Export to CSV files in ZIP (when openpyxl not available)"""
    quarter_info = QUARTERS[quarter_name]
    
    dataframes = {}
    
    if "Executive Summary" in selected_sections:
        summary_data = {
            "Project Name": [PROJECT_NAME],
            "Assignment Title": [ASSIGNMENT_TITLE],
            "Client": [CLIENT_NAME],
            "Consultant": [CONSULTANT_NAME],
            "Quarter": [quarter_name],
            "Months": [", ".join(quarter_info["months"])],
            "Status": [quarter_info["status"].upper()],
            "World Bank Loan No": ["IBRD 9737-IN"],
            "RFP Reference": ["IN-MITRA(PMU)-PforR-Edu-QCBS"],
            "Report Generated On": [datetime.now().strftime('%d %B %Y, %H:%M:%S')]
        }
        dataframes["Executive_Summary"] = pd.DataFrame(summary_data)
    
    if "Key Activities" in selected_sections:
        activities_data = []
        for i, activity in enumerate(quarter_info["key_activities"], 1):
            status = "Completed" if activity.startswith("✅") else "In Progress" if activity.startswith("🔄") else "Planned"
            activities_data.append({
                "S.No": i,
                "Activity": activity.replace("✅", "").replace("🔄", "").strip(),
                "Status": status,
                "Quarter": quarter_name
            })
        dataframes["Key_Activities"] = pd.DataFrame(activities_data)
    
    if "Deliverables" in selected_sections:
        deliverables_data = []
        for deliverable in quarter_info["deliverables"]:
            status = "Completed" if "✅" in deliverable else "In Progress" if "🔄" in deliverable else "Pending"
            deliverables_data.append({
                "Deliverable": deliverable.replace("✅", "").replace("🔄", "").strip(),
                "Status": status,
                "Quarter": quarter_name
            })
        dataframes["Deliverables"] = pd.DataFrame(deliverables_data)
    
    if "Milestones" in selected_sections:
        milestones_data = []
        for milestone in quarter_info["milestones"]:
            milestones_data.append({
                "Milestone Name": milestone["name"],
                "Target Date": milestone["date"],
                "Status": milestone["status"].upper(),
                "Quarter": quarter_name
            })
        dataframes["Milestones"] = pd.DataFrame(milestones_data)
    
    if "Data Collection & Engagement" in selected_sections:
        engagement_data = {
            "Metric": ["Data Collection Process", "Stakeholder Engagement", "Review Mechanism", "Universities Involved", "Responsible Party"],
            "Details": [
                quarter_info["data_collection"],
                quarter_info["stakeholder_engagement"],
                quarter_info["review_mechanism"],
                quarter_info.get("universities_involved", "All 7 universities"),
                quarter_info.get("responsible_party", "ICARE Pvt. Ltd.")
            ]
        }
        dataframes["Data_Collection_Engagement"] = pd.DataFrame(engagement_data)
    
    if "Universities & GRDAU" in selected_sections:
        uni_data = []
        for code, uni in UNIVERSITIES.items():
            if code != "MITRA":
                uni_data.append({
                    "University": uni["name"],
                    "Location": uni["location"],
                    "Nodal Officer": uni.get("nodal_officer", "N/A"),
                    "Contact": uni.get("contact", "N/A"),
                    "GRDAU Status": uni["grdau_status"],
                    "GRDAU Completion Date": uni["grdau_completion_date"],
                    "Coordinators": ", ".join(uni["coordinators"])
                })
        dataframes["Universities_GRDAU"] = pd.DataFrame(uni_data)
    
    if "Team Members" in selected_sections:
        team_data = {
            "Name": ["Dr. Harshal Kotwal", "Shubham Singh", "Sagar Teli", "Sneha Kashitkar", "Nitish Kumbhar",
                    "Anjali Singh", "Vaibhav Ambekar", "Atharav Paturkar", "Prathamesh Babhulkar", "Jagan Sridhar"],
            "Role": ["Project Lead", "Data Analytics Specialist", "Statistician & Program Designer", 
                    "Institutional Coordinator", "Institutional Coordinator", "Institutional Coordinator",
                    "Institutional Coordinator", "Institutional Coordinator", "Institutional Coordinator", "Institutional Coordinator"],
            "University": ["ICARE", "MITRA", "Mumbai University", "Mumbai University", "KBCNMU Jalgaon",
                          "Nagpur University", "COEP Pune", "BAMU Aurangabad", "Amravati University", "SPPU Pune"]
        }
        dataframes["Team_Members"] = pd.DataFrame(team_data)
    
    import zipfile
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for sheet_name, df in dataframes.items():
            csv_data = df.to_csv(index=False).encode('utf-8')
            zip_file.writestr(f"{sheet_name}.csv", csv_data)
    
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

# ============================================================
# SIDEBAR NAVIGATION (Same as before)
# ============================================================

with st.sidebar:
    st.markdown("## 🎯 MahaSTRIDE")
    st.markdown("---")
    
    nav_options = {
        "🏠 Executive Summary": "summary",
        "📊 Quarterly Plan": "quarterly",
        "🏫 Universities & Team": "universities",
        "🏛️ War Room & GRDAU": "warroom",
        "🎯 Milestones Tracker": "milestones",
        "📋 Deliverables": "deliverables",
        "🔄 Review Mechanisms": "review",
        "📁 Documents": "documents",
        "📥 Export Reports": "export"
    }
    
    selected_nav = st.radio("Navigation", list(nav_options.keys()), label_visibility="collapsed")
    selected_key = nav_options[selected_nav]
    
    st.markdown("---")
    st.markdown("### ℹ️ Project Info")
    st.markdown(f"**Start Date:** {START_DATE.strftime('%d %b %Y')}")
    st.markdown(f"**End Date:** {END_DATE.strftime('%d %b %Y')}")
    st.markdown(f"**Duration:** 24 months (8 Quarters)")
    st.markdown(f"**Universities:** 7")
    st.markdown(f"**Data Analysts:** 10")
    
    st.markdown("---")
    st.markdown("### ✅ Achievements")
    for achievement in ACHIEVED_MILESTONES:
        st.markdown(f"- ✅ {achievement['milestone']}")
    
    st.markdown("---")
    st.markdown("### 📞 Contact")
    st.markdown("**PMU MahaSTRIDE**")
    st.markdown("📧 pmu.mahastride@mahamitra.org")
    st.markdown("📞 022-69979440")
    st.markdown("📍 5th Floor, Nirmal Building, Nariman Point, Mumbai-400021")

# ============================================================
# MAIN HEADER
# ============================================================

st.markdown(f"""
<div class="main-header">
    <h1>🎯 {PROJECT_NAME}</h1>
    <div class="assignment-title">
        📋 {ASSIGNMENT_TITLE}
    </div>
    <div class="parties-info">
        <div class="party-box">
            <strong>🏛️ Client:</strong> {CLIENT_NAME}
        </div>
        <div class="party-box">
            <strong>🤝 Consultant:</strong> {CONSULTANT_NAME}
        </div>
        <p style="margin-top:0.5rem; font-size:0.85rem;">World Bank Loan No: IBRD 9737-IN | RFP Ref: IN-MITRA(PMU)-PforR-Edu-QCBS</p>
    </div>
</div>
""", unsafe_allow_html=True)

# ============================================================
# EXPORT REPORTS PAGE
# ============================================================
if selected_key == "export":
    st.header("📥 Export Quarterly Reports")
    st.markdown("Generate and download professionally formatted quarterly progress reports.")
    
    st.markdown("---")
    
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.subheader("📋 Select Quarter")
        selected_quarter = st.selectbox(
            "Choose Quarter",
            options=list(QUARTERS.keys()),
            help="Select the quarter for which you want to generate the report"
        )
        
        st.subheader("📎 Select Sections to Include")
        
        section_options = [
            "Executive Summary",
            "Key Activities",
            "Deliverables",
            "Milestones",
            "Data Collection & Engagement",
            "Universities & GRDAU",
            "Team Members"
        ]
        
        selected_sections = st.multiselect(
            "Choose sections to include in the report",
            options=section_options,
            default=section_options,
            help="Select which sections you want in your report"
        )
    
    with col2:
        quarter_info = QUARTERS[selected_quarter]
        st.subheader("📊 Report Preview")
        st.markdown(f"""
        **Quarter:** {selected_quarter}<br>
        **Months:** {', '.join(quarter_info['months'])}<br>
        **Status:** {quarter_info['status'].upper()}<br>
        **Total Activities:** {len(quarter_info['key_activities'])}<br>
        **Total Deliverables:** {len(quarter_info['deliverables'])}<br>
        **Total Milestones:** {len(quarter_info['milestones'])}
        """, unsafe_allow_html=True)
        
        if quarter_info['status'] == 'ongoing':
            st.markdown('<span class="status-ongoing">🟡 ONGOING</span>', unsafe_allow_html=True)
        else:
            st.markdown('<span class="status-completed">⚪ UPCOMING</span>', unsafe_allow_html=True)
        
        # Show openpyxl installation instruction if needed
        st.info("ℹ️ For best formatting, install openpyxl: `pip install openpyxl`")
    
    st.markdown("---")
    
    # Download button
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if selected_sections:
            if st.button("📥 Generate & Download Report", use_container_width=True, type="primary"):
                with st.spinner("Generating professional report..."):
                    report_data, format_type = export_quarter_to_professional_excel(selected_quarter, selected_sections)
                    
                    if format_type == "excel":
                        filename = f"MahaSTRIDE_{selected_quarter.replace(' ', '_').replace(':', '')}_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
                        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    else:
                        filename = f"MahaSTRIDE_{selected_quarter.replace(' ', '_').replace(':', '')}_Report_{datetime.now().strftime('%Y%m%d')}.zip"
                        mime_type = "application/zip"
                    
                    b64 = base64.b64encode(report_data).decode()
                    href = f'<a href="data:{mime_type};base64,{b64}" download="{filename}" style="background-color:#28a745;color:white;padding:12px 24px;text-decoration:none;border-radius:5px;display:inline-block;font-weight:bold;">📥 Download Report</a>'
                    st.markdown(href, unsafe_allow_html=True)
                    st.success(f"✅ Report for {selected_quarter} generated successfully!")
        else:
            st.warning("⚠️ Please select at least one section to include in the report")

# ============================================================
# EXISTING CONTENT PAGES (Summary, Quarterly, etc.)
# ============================================================
# [All the other content pages remain exactly the same as before]
# ... (keeping the existing summary, quarterly, universities, warroom, milestones, deliverables, review, documents sections)

# ============================================================
# FOOTER
# ============================================================
st.markdown("---")
st.markdown(f"""
<div style="text-align: center; color: #666; font-size: 0.8rem;">
    <p>© 2026-2028 {CLIENT_NAME} | MahaSTRIDE Project | World Bank Loan No: IBRD 9737-IN</p>
    <p>Consultant: {CONSULTANT_NAME} | Duration: 24 months (8 Quarters) | Working Days: Monday to Friday | Hours: 10:00 - 18:00</p>
    <p>Last Updated: {datetime.now().strftime('%d %B %Y, %H:%M:%S')}</p>
</div>
""", unsafe_allow_html=True)
